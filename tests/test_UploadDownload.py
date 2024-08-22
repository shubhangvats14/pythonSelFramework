from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl

chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
service_obj = Service(r"C:\Users\shubhangvats\OneDrive - Hexaview Technologies\Downloads\chromedriver-win64"
                      r"\chromedriver-win64\chromedriver.exe")
driver = webdriver.Chrome(service=service_obj, options=chrome_options)
driver.get("https://rahulshettyacademy.com/upload-download-test")
driver.implicitly_wait(5)

# ###### DOWNLOADING excel File
driver.find_element(By.ID, "downloadButton").click()

# ##### EDITING the downloaded Excel file
file_path = "C:/Users/shubhangvats/OneDrive - Hexaview Technologies/Downloads/download.xlsx"


def update_cell_value(file_p, item_name, column_n, new_value):
    Dict = {}
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    # getting Price column index
    for c in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=c).value == column_n:
            Dict["col"] = c

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == item_name:
                Dict["row"] = i
            else:
                continue
    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_p)
    book.close()


update_cell_value(file_path, "Apple", "price", "999")

# ##### UPLOADING the edited Excel file
upload_btn = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
upload_btn.send_keys(file_path)
# if the locator has input tag with type attribute = file, then we can send the file location using sen_keys method
# and that will upload the file

toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
column_name = 'Price'
fruit = 'Apple'
WebDriverWait(driver, 5).until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
column = driver.find_element(By.XPATH, "//div[text()='" + column_name + "']").get_attribute("data-column-id")
# Here we have used the column_name variable to generalize the column in order to get the locator for any column
# when the locator is reached to the column, we brought the value of "data-column-id", which is the index of the row

actual_price = driver.find_element(By.XPATH, "//div[text()= '" + fruit + "']/parent::div/parent::div/"
                                                                         "div[@id='cell-" + column + "-undefined']").text
# Here we are using the parent traversal, as we traversed from Apple to the parent div(s) togs
# Now, using the fruit name and the column index, we brought the value on the cell

print(actual_price)

driver.close()
