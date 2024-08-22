# #### Understanding the openpyexcel commands to manipulate Excel files #########

import openpyxl  # used to get and write back excel data

dict_data = {}
book = openpyxl.load_workbook("TestData.xlsx")  # workbook path
sheet = book.active  # get the active sheet, which is the first one
cell = sheet.cell(row=1, column=2)  # get a cell value using row, column
print(cell.value)
sheet.cell(row=2, column=2).value = "Rahul"  # write back a cell value using row, column
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)  # max rows containing values
print(sheet.max_column)  # max columns containing values

print(sheet['B5'].value)  # simpler way to get a cell value
sheet['B5'].value = "New_value"  # simpler way to write a cell value
print(sheet['B5'].value)

book.save("TestData.xlsx")

# To print all values from sheet in same format
print("\n")
for r in range(1, sheet.max_row + 1):  # iterating rows
    if sheet.cell(row=r, column=1).value == "test_case3":  # for getting data of just test_case3
        for c in range(1, sheet.max_column + 1):  # iterating columns
            print(sheet.cell(row=r, column=c).value, end=' ')  # end to print horizontally with ' ' between them
        print("\n")

# To save the test case data in dictionary
for r in range(1, sheet.max_row + 1):  # iterating rows
    if sheet.cell(row=r, column=1).value == "test_case3":  # for getting data of just test_case3
        for c in range(2, sheet.max_column + 1):
            # dict_data["last_name"] = "Vats"
            dict_data[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value

print(dict_data)
