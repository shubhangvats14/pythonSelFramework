import openpyxl


class HomePageData:
    # home_page_data = [{"name": "Chrome", "email": "Shubhang@fhg.com", "password": "Vats"},
    #                   {"name": "Firefox", "email": "Rahul@ghg.com", "password": "Shetty"},
    #                   {"name": "IE", "email": "Vijay@ghvh.com", "password": "Rana"}]

    # Now we read the above data from an Excel file
    @staticmethod
    def getTestData(test_case_name):  # test_case_name will come from our test
        dict_data = {}
        book = openpyxl.load_workbook(r"C:\Users\shubhangvats\PycharmProjects\pythonSelFramework"
                                      r"\TestData\TestData.xlsx")
        sheet = book.active
        for r in range(1, sheet.max_row + 1):  # iterating rows
            if sheet.cell(row=r, column=1).value == test_case_name:  # for getting data of just test_case3
                for c in range(2, sheet.max_column + 1):
                    # dict_data["last_name"] = "Vats"
                    dict_data[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value
        return [dict_data]
